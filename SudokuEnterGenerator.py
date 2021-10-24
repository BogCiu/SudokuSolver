from openpyxl.styles.borders import Border, Side
import openpyxl as xl

sudoku = xl.Workbook()
sheet = sudoku['Sheet']
for i in range(1, 10):
    for j in range(1,10):
        if i % 3 == 1 and j % 3 == 1:
            sheet.cell(i, j).border = Border(left=Side(style='thick'),
                                             top=Side(style='thick'))
        elif i % 3 == 1 and j % 3 == 2:
            sheet.cell(i, j).border = Border(top=Side(style='thick'))
        elif i % 3 == 1 and j % 3 == 0:
            sheet.cell(i, j).border = Border(right=Side(style='thick'),
                                             top=Side(style='thick'))

        if i % 3 == 2 and j % 3 == 1:
            sheet.cell(i, j).border = Border(left=Side(style='thick'))
#skip i%2 and j%2, has no borders
        elif i % 3 == 2 and j % 3 == 0:
            sheet.cell(i, j).border = Border(right=Side(style='thick'))

        if i % 3 == 0 and j % 3 == 1:
            sheet.cell(i, j).border = Border(left=Side(style='thick'),
                                             bottom=Side(style='thick'))
        elif i % 3 == 0 and j % 3 == 2:
            sheet.cell(i, j).border = Border(bottom=Side(style='thick'))
        elif i % 3 == 0 and j % 3 == 0:
            sheet.cell(i, j).border = Border(right=Side(style='thick'),
                                             bottom=Side(style='thick'))

for i in range(1,10):
    sheet.column_dimensions[xl.utils.get_column_letter(i)].width = 5;

sudoku.save("SudokuExcel.xlsx")
