import openpyxl as xl
import numpy
import SudokuCheckFunctions as Check

is_valid = False
is_finished = True
step_forward = True


def sudoku_step(co, ro, forward):
    logic_error_variable = False
    if forward:
        if co < 8:
            co += 1
        elif ro < 8:
            co = 0
            ro += 1
        else:
            logic_error_variable = True
            print("Exiting due to error: moving forward from cell 9-9")
    else:
        if co > 0:
            co -= 1
        elif ro > 0:
            co = 8
            ro -= 1
        else:
            logic_error_variable = True
            print("Exiting due to error: trying to traceback from cell 1-1")
    return [co, ro, logic_error_variable]


class SudokuCell:
    is_given = None
    last_attempted_value = 0

    def __init__(self, r_nr, c_nr):
        self.row = r_nr
        self.column = c_nr
        self.current_value = (row * 10 + column)


c = numpy.empty((9, 9), dtype=object)

wb = xl.open("SudokuExcel.xlsx")
ws = wb['Sheet']

for row in range(1, 10):
    for column in range(1, 10):
        c[row - 1][column - 1] = SudokuCell(row, column)
        if ws.cell(row, column).value:
            c[row - 1][column - 1].is_given = True
            c[row - 1][column - 1].current_value = ws.cell(row, column).value
        else:
            c[row - 1][column - 1].is_given = False
#            print (f'cell c{row-1},{column-1} contains :{c[row - 1][column - 1].current_value}') TESTING VISUALISATION, DELETE AFTER FINAL

solver_row = 0
solver_column = 0
solver_forward = True
logic_error = False

is_valid = Check.check_validity(c)

is_finished = Check.check_finished(c)

while not is_finished or not is_valid:
    if logic_error:
        break
    if not c[solver_row][solver_column].is_given:
        if c[solver_row][solver_column].last_attempted_value < 9:
            c[solver_row][solver_column].current_value = c[solver_row][solver_column].last_attempted_value + 1
            c[solver_row][solver_column].last_attempted_value = c[solver_row][solver_column].current_value
            is_valid = Check.check_validity(c)
            is_finished = Check.check_finished(c)
            if is_valid and not is_finished:
                solver_forward = True
                [solver_column, solver_row, logic_error] = sudoku_step(solver_column, solver_row, solver_forward)
                # check for finished as well?
        else:
            c[solver_row][solver_column].current_value = (solver_row + 1)*10 + solver_column + 1
            c[solver_row][solver_column].last_attempted_value = 0
            solver_forward = False
            [solver_column, solver_row, logic_error] = sudoku_step(solver_column, solver_row, solver_forward)
    else:
        [solver_column, solver_row, logic_error] = sudoku_step(solver_column, solver_row, solver_forward)

# for i in range(0, 9):
#   for j in range(0, 9):
#       print(c[i][j].current_value, end="")
#   print("\n")

for row in range(1, 10):
    for column in range(1, 10):
        ws.cell(row, column).value = c[row - 1][column - 1].current_value

wb.save("SudokuExcelSolved.xlsx")

is_valid = Check.check_validity(c)
print(f'Puzzle validity is {is_valid}')

is_finished = Check.check_finished(c)
print(f'Puzzle finished status is {is_finished}')

